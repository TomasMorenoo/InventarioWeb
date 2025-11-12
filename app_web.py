"""
app_web_completo.py - Dashboard Web COMPLETAMENTE INTEGRADO con estados
Incluye soporte completo para MAC Address + campo estado + estructura unificada
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, make_response
import sqlite3
import os
from datetime import datetime
import csv
from io import StringIO
import logging
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import get_config

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

app = Flask(__name__)
config_obj = get_config()
app.config.from_object(config_obj)
config_obj.init_app(app)

class UnifiedWebDatabaseManager:
    """Gestor de BD web COMPLETAMENTE UNIFICADO con estados"""
    
    def __init__(self):
        self.inventario_db_path = self._get_active_inventario_db()
        self.oficinas_db_path = app.config['OFICINAS_CNE_DB']
        self.impresoras_db_path = self._get_active_impresoras_db()
        self.notebooks_db_path = self._get_active_notebooks_db()  # NUEVO
        self.is_shared = self._is_shared_database()
        self._init_unified_inventario_db()
        self._init_impresoras_db()
        self._init_notebooks_db()  # NUEVO
        
    def _get_active_inventario_db(self):
        """Determina qué BD de inventario usar con fallback"""
        try:
            shared_dir = os.path.dirname(app.config['INVENTARIO_DB'])
            if shared_dir and not os.path.exists(shared_dir):
                try:
                    os.makedirs(shared_dir, exist_ok=True)
                except Exception:
                    pass
            
            # Intenta conectar a la BD compartida
            test_conn = sqlite3.connect(app.config['INVENTARIO_DB'], timeout=5)
            test_conn.execute("SELECT 1")
            test_conn.close()
            print(f"✅ Usando BD compartida: {app.config['INVENTARIO_DB']}")
            return app.config['INVENTARIO_DB']
            
        except Exception as e:
            print(f"❌ No se puede acceder a BD compartida: {e}")
            print(f"✅ Usando BD local: {app.config['LOCAL_FALLBACK_DB']}")
            return app.config['LOCAL_FALLBACK_DB']
    
    def _get_active_impresoras_db(self):
        """Determina qué BD de impresoras usar con fallback"""
        try:
            shared_dir = os.path.dirname(app.config['IMPRESORAS_DB'])
            if shared_dir and not os.path.exists(shared_dir):
                try:
                    os.makedirs(shared_dir, exist_ok=True)
                except Exception:
                    pass
            
            test_conn = sqlite3.connect(app.config['IMPRESORAS_DB'], timeout=5)
            test_conn.execute("SELECT 1")
            test_conn.close()
            print(f"✅ Usando BD impresoras compartida: {app.config['IMPRESORAS_DB']}")
            return app.config['IMPRESORAS_DB']
            
        except Exception as e:
            print(f"❌ BD impresoras compartida no : {e}")
            print(f"✅ Usando BD impresoras local: {app.config['LOCAL_IMPRESORAS_DB']}")
            return app.config['LOCAL_IMPRESORAS_DB']
    
    def _get_active_notebooks_db(self):
        """Determina qué BD de notebooks usar con fallback"""
        try:
            shared_dir = os.path.dirname(app.config['NOTEBOOKS_DB'])
            if shared_dir and not os.path.exists(shared_dir):
                try:
                    os.makedirs(shared_dir, exist_ok=True)
                except Exception:
                    pass
            
            test_conn = sqlite3.connect(app.config['NOTEBOOKS_DB'], timeout=5)
            test_conn.execute("SELECT 1")
            test_conn.close()
            print(f"✅ Usando BD notebooks compartida: {app.config['NOTEBOOKS_DB']}")
            return app.config['NOTEBOOKS_DB']
            
        except Exception as e:
            print(f"❌ BD notebooks compartida no disponible: {e}")
            print(f"✅ Usando BD notebooks local: {app.config['LOCAL_NOTEBOOKS_DB']}")
            return app.config['LOCAL_NOTEBOOKS_DB']
    
    def _is_shared_database(self):
        """Determina si estamos usando las BD compartidas"""
        return "16.1.1.118" in self.inventario_db_path
    
    def _init_unified_inventario_db(self):
        """Inicializa la tabla de inventario CON ESTRUCTURA COMPLETA incluyendo estado"""
        try:
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                
                # Crear tabla con TODAS las columnas incluyendo estado
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS inventario (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        usuario TEXT,
                        oficina_id INTEGER,
                        pc_nombre TEXT,
                        pc_usuario TEXT,
                        windows TEXT,
                        ip TEXT,
                        marca TEXT,
                        modelo TEXT,
                        numero_serie TEXT,
                        mac_address TEXT,
                        usa_ocs INTEGER DEFAULT 0,
                        contrasena TEXT,
                        procesador TEXT,
                        ram TEXT,
                        disco TEXT,
                        motherboard TEXT,
                        tarjeta_grafica TEXT,
                        estado TEXT DEFAULT 'Ok',
                        fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        sync_status TEXT DEFAULT 'synced',
                        observaciones TEXT
                    )
                """)
                
                # Verificar si necesitamos agregar columnas faltantes
                cursor.execute("PRAGMA table_info(inventario)")
                columnas_existentes = [col[1] for col in cursor.fetchall()]
                
                columnas_requeridas = [
                    'mac_address', 'procesador', 'ram', 'disco', 
                    'motherboard', 'tarjeta_grafica', 'estado', 'piso', 'sync_status', 'observaciones'
                ]
                
                for columna in columnas_requeridas:
                    if columna not in columnas_existentes:
                        if columna == 'mac_address':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN mac_address TEXT")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                        elif columna == 'procesador':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN procesador TEXT")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                        elif columna == 'ram':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN ram TEXT")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                        elif columna == 'disco':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN disco TEXT")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                        elif columna == 'motherboard':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN motherboard TEXT")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                        elif columna == 'tarjeta_grafica':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN tarjeta_grafica TEXT")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                        elif columna == 'estado':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN estado TEXT DEFAULT 'Ok'")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                        elif columna == 'piso':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN piso TEXT")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                        elif columna == 'sync_status':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN sync_status TEXT DEFAULT 'synced'")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                        elif columna == 'observaciones':
                            cursor.execute("ALTER TABLE inventario ADD COLUMN observaciones TEXT")
                            print(f"✅ Columna '{columna}' agregada a la tabla inventario")
                
                conn.commit()
                print("✅ Tabla inventario inicializada/actualizada con estructura completa")
                
        except Exception as e:
            print(f"❌ Error inicializando BD inventario: {e}")
    
    def get_inventario_connection(self):
        """Conexión a BD de inventario con reintentos"""
        try:
            conn = sqlite3.connect(self.inventario_db_path, timeout=10)
            conn.execute("SELECT 1")
            return conn
        except Exception:
            # Si falla la BD compartida, usar local
            if "16.1.1.118" in self.inventario_db_path:
                self.inventario_db_path = app.config['LOCAL_FALLBACK_DB']
                self._init_unified_inventario_db()  # Inicializar BD local
            return sqlite3.connect(self.inventario_db_path, timeout=10)
    
    def get_oficinas_connection(self):
        """Conexión a BD de oficinas CNE"""
        try:
            return sqlite3.connect(self.oficinas_db_path, timeout=5)
        except Exception:
            # Si no hay BD de oficinas, crear una local básica
            return self._create_local_oficinas_db()
    
    def _create_local_oficinas_db(self):
        """Crea una BD local de oficinas si no existe la principal"""
        local_oficinas = "oficinas_local.db"
        conn = sqlite3.connect(local_oficinas)
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS oficinas (
                id INTEGER PRIMARY KEY,
                nombre_oficina TEXT,
                piso INTEGER
            )
        """)
        
        # Oficinas por defecto
        oficinas_default = [
            (1, "Oficina Principal", 1),
            (2, "Oficina Secundaria", 2),
            (3, "Oficina Norte", 1),
            (4, "Oficina Sur", 3)
        ]
        
        cursor.executemany("INSERT OR IGNORE INTO oficinas VALUES (?, ?, ?)", oficinas_default)
        conn.commit()
        
        return conn
    
    def get_impresoras_connection(self):
        """Conexión a BD de impresoras"""
        try:
            conn = sqlite3.connect(self.impresoras_db_path, timeout=5)
            conn.execute("SELECT 1")
            return conn
        except Exception:
            if "16.1.1.118" in self.impresoras_db_path:
                self.impresoras_db_path = app.config['LOCAL_IMPRESORAS_DB']
                self._init_impresoras_db()
            return sqlite3.connect(self.impresoras_db_path)
    
    def is_record_duplicate(self, inventory_id):
        """Verifica si un registro específico es duplicado"""
        try:
            registro = self.get_inventory_by_id(inventory_id)
            if not registro:
                return False    
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar duplicados por serie
                if registro['serie']:
                    cursor.execute("SELECT COUNT(*) FROM inventario WHERE numero_serie = ? AND numero_serie != ''", (registro['serie'],))
                    if cursor.fetchone()[0] > 1:
                        return True
                
                # Verificar duplicados por IP
                if registro['ip_pc']:
                    cursor.execute("SELECT COUNT(*) FROM inventario WHERE ip = ? AND ip != ''", (registro['ip_pc'],))
                    if cursor.fetchone()[0] > 1:
                        return True
                
                # Verificar duplicados por MAC
                if registro['mac_address']:
                    cursor.execute("SELECT COUNT(*) FROM inventario WHERE mac_address = ? AND mac_address != ''", (registro['mac_address'],))
                    if cursor.fetchone()[0] > 1:
                        return True
                
                # Verificar duplicados por nombre PC
                if registro['nombre_pc']:
                    cursor.execute("SELECT COUNT(*) FROM inventario WHERE pc_nombre = ? AND pc_nombre != ''", (registro['nombre_pc'],))
                    if cursor.fetchone()[0] > 1:
                        return True
                        
                return False
                
        except Exception as e:
            print(f"Error verificando duplicado para registro {inventory_id}: {e}")
            return False
    
    def _check_duplicate_direct(self, current_id, serie, ip, mac, nombre_pc):
        """Verifica duplicados sin llamar get_inventory_by_id para evitar bucle infinito"""
        try:
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar duplicados por serie (excluyendo el registro actual)
                if serie:
                    cursor.execute("SELECT COUNT(*) FROM inventario WHERE numero_serie = ? AND numero_serie != '' AND id != ?", (serie, current_id))
                    if cursor.fetchone()[0] > 0:
                        return True
                
                # Verificar duplicados por IP
                if ip:
                    cursor.execute("SELECT COUNT(*) FROM inventario WHERE ip = ? AND ip != '' AND id != ?", (ip, current_id))
                    if cursor.fetchone()[0] > 0:
                        return True
                
                # Verificar duplicados por MAC
                if mac:
                    cursor.execute("SELECT COUNT(*) FROM inventario WHERE mac_address = ? AND mac_address != '' AND id != ?", (mac, current_id))
                    if cursor.fetchone()[0] > 0:
                        return True
                
                # Verificar duplicados por nombre PC
                if nombre_pc:
                    cursor.execute("SELECT COUNT(*) FROM inventario WHERE pc_nombre = ? AND pc_nombre != '' AND id != ?", (nombre_pc, current_id))
                    if cursor.fetchone()[0] > 0:
                        return True
                        
                return False
                
        except Exception as e:
            print(f"Error verificando duplicado: {e}")
            return False

    def get_all_inventory(self):
        """Obtiene todo el inventario CON CAMPOS REALES DE LA BD INCLUYENDO ESTADO"""
        try:
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                # ✅ CORREGIDO: Incluye inv.estado en SELECT
                cursor.execute("""
                    SELECT inv.id, inv.usuario, inv.oficina_id, inv.pc_nombre, inv.pc_usuario, 
                        inv.windows, inv.ip, inv.marca, inv.modelo, inv.numero_serie, 
                        inv.mac_address, inv.usa_ocs, inv.contrasena, inv.procesador,
                        inv.ram, inv.disco, inv.estado, inv.piso, inv.fecha_registro, inv.fecha_modificacion
                    FROM inventario inv
                    ORDER BY inv.pc_nombre
                """)
                
                registros_raw = cursor.fetchall()
                registros = []
                
                for row in registros_raw:
                    try:
                        # ✅ CORREGIDO: Incluye estado_bd en unpacking
                        (id_inv, usuario, oficina_id, pc_nombre, pc_usuario, windows, ip, 
                        marca, modelo, numero_serie, mac_address, usa_ocs, contrasena, 
                        procesador, ram, disco, estado_bd, piso_bd, fecha_registro, fecha_modificacion,
                        oficina_nombre, oficina_piso) = row
                    except ValueError:
                        # Si faltan columnas, rellenar con valores por defecto
                        # ✅ CORREGIDO: Aumentado a 22 columnas para incluir estado
                        row_data = list(row) + [None] * (22 - len(row))
                        (id_inv, usuario, oficina_id, pc_nombre, pc_usuario, windows, ip, 
                        marca, modelo, numero_serie, mac_address, usa_ocs, contrasena, 
                        procesador, ram, disco, estado_bd, piso_bd, fecha_registro, fecha_modificacion,
                        oficina_nombre, oficina_piso) = row_data[:22]
                    
                    if not oficina_nombre and oficina_id:
                        oficina_nombre = self._get_oficina_name_from_cne(oficina_id)
                    
                    registros.append({
                        'id': id_inv,
                        'tecnico': usuario or '',
                        'oficina': oficina_nombre or f'ID: {oficina_id}',
                        'piso': piso_bd if piso_bd else (str(oficina_piso) if oficina_piso is not None else ''),
                        'usuario_persona': usuario or '',       # ✅ CORRECCIÓN: Persona = usuario
                        'pc_usuario': pc_usuario or '',         # ✅ CORRECCIÓN: Usuario red = pc_usuario
                        'nombre_pc': pc_nombre or '',
                        'ip_pc': ip or '',
                        'mac_address': mac_address or '',
                        'marca': marca or '',
                        'modelo': modelo or '',
                        'serie': numero_serie or '',
                        'procesador': procesador or '',
                        'ram': ram or '',
                        'disco': disco or '',
                        'motherboard': '',
                        'tarjeta_grafica': '',
                        'windows': windows or '',
                        'estado': estado_bd or 'Ok',  # ✅ CORREGIDO: Lee estado real de BD
                        'fecha_creacion': fecha_registro or '',
                        'fecha_modificacion': fecha_modificacion or '',
                        'usa_ocs': usa_ocs,
                        'contrasena': contrasena or '',
                        'sync_status': 'synced',
                        'observaciones': '',
                        'is_duplicate': 0
                    })
                
                return registros
                
        except Exception as e:
            print(f"Error obteniendo inventario: {e}")
            return []
    
    def get_inventory_by_id(self, inventory_id):
        """Obtiene un registro específico por ID CON CAMPOS REALES DE LA BD INCLUYENDO ESTADO"""
        try:
            if inventory_id == 16:
                print(f"🔍 DEBUG REGISTRO 16: Leyendo registro...")
                
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                
                # DEBUG: Verificar oficina_id real en inventario
                if inventory_id == 16:
                    cursor.execute("SELECT oficina_id FROM inventario WHERE id = 16")
                    oficina_id_real = cursor.fetchone()
                    print(f"🔍 DEBUG REGISTRO 16: oficina_id real en inventario: {oficina_id_real}")
                
                cursor.execute("""
                    SELECT inv.id, inv.usuario, inv.oficina_id, inv.pc_nombre, inv.pc_usuario, 
                        inv.windows, inv.ip, inv.marca, inv.modelo, inv.numero_serie, 
                        inv.mac_address, inv.usa_ocs, inv.contrasena, inv.procesador,
                        inv.ram, inv.disco, inv.estado, inv.piso, inv.fecha_registro, inv.fecha_modificacion
                    FROM inventario inv
                    WHERE inv.id = ?
                """, (inventory_id,))

                row = cursor.fetchone()
                if row:
                    # Desempaqueta SIN los datos de oficina del JOIN
                    (id_inv, usuario, oficina_id, pc_nombre, pc_usuario, windows, ip, 
                    marca, modelo, numero_serie, mac_address, usa_ocs, contrasena, 
                    procesador, ram, disco, estado_bd, piso_bd, fecha_registro, fecha_modificacion) = row
                    
                    # Obtén el nombre de oficina usando la conexión correcta
                    oficina_nombre = None
                    oficina_piso_cne = None
                    if oficina_id:
                        oficina_info = self.get_oficina_with_piso_by_id(oficina_id)
                        if oficina_info:
                            oficina_nombre = oficina_info.get('nombre')
                            oficina_piso_cne = oficina_info.get('piso')
                    
                    if inventory_id == 16:
                        print(f"🔍 DEBUG REGISTRO 16: oficina_nombre obtenida: '{oficina_nombre}'")
                        print(f"🔍 DEBUG REGISTRO 16: oficina_piso_cne: '{oficina_piso_cne}'")
                    
                    return {
                        'id': id_inv,
                        'tecnico': usuario or '',
                        'oficina': oficina_nombre or f'ID: {oficina_id}',
                        'piso': piso_bd if piso_bd else (str(oficina_piso_cne) if oficina_piso_cne is not None else ''),
                        'usuario_persona': usuario or '',
                        'pc_usuario': pc_usuario or '',
                        'nombre_pc': pc_nombre or '',
                        'ip_pc': ip or '',
                        'mac_address': mac_address or '',
                        'marca': marca or '',
                        'modelo': modelo or '',
                        'serie': numero_serie or '',
                        'procesador': procesador or '',
                        'ram': ram or '',
                        'disco': disco or '',
                        'motherboard': '',
                        'tarjeta_grafica': '',
                        'windows': windows or '',
                        'estado': estado_bd or 'Ok',
                        'fecha_creacion': fecha_registro or '',
                        'fecha_modificacion': fecha_modificacion or '',
                        'usa_ocs': usa_ocs,
                        'contrasena': contrasena or '',
                        'sync_status': 'synced',
                        'observaciones': '',
                        'is_duplicate': 1 if self._check_duplicate_direct(id_inv, numero_serie, ip, mac_address, pc_nombre) else 0
                    }
            return None
        except Exception as e:
            print(f"Error obteniendo registro {inventory_id}: {e}")
            return None
        
        

    def get_oficina_with_piso_by_id(self, oficina_id):
        """Obtiene oficina por ID desde BD CNE"""
        try:
            with self.get_oficinas_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, nombre_oficina, piso FROM oficinas WHERE id = ?", (oficina_id,))
                result = cursor.fetchone()
                if result:
                    return {'id': result[0], 'nombre': result[1], 'piso': result[2]}
                return None
        except Exception:
            return None
    
    def update_inventory_record(self, inventory_id, data):
        """Actualiza un registro CON TODOS LOS CAMPOS incluyendo estado y piso"""
        try:
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                
                # Buscar oficina_id
                oficina_nombre = data.get('oficina', '')
                oficina_id = None
                print(f"🔍 DEBUG: Buscando oficina: '{oficina_nombre}'")
                if oficina_nombre:
                    try:
                        with self.get_oficinas_connection() as ofi_conn:
                            ofi_cursor = ofi_conn.cursor()
                            ofi_cursor.execute("SELECT id, nombre_oficina FROM oficinas WHERE nombre_oficina = ?", (oficina_nombre,))
                            result = ofi_cursor.fetchone()
                            if result:
                                oficina_id = result[0]
                                print(f"✅ DEBUG: Oficina encontrada - ID: {oficina_id}, Nombre: {result[1]}")
                            else:
                                print(f"❌ DEBUG: Oficina '{oficina_nombre}' NO encontrada en BD oficinas")
                                # Listar todas las oficinas para debug
                                ofi_cursor.execute("SELECT id, nombre_oficina FROM oficinas LIMIT 10")
                                todas = ofi_cursor.fetchall()
                                print(f"🔎 DEBUG: Oficinas disponibles: {todas}")
                    except Exception as e:
                        print(f"❌ DEBUG: Error conectando a BD oficinas: {e}")
                
                # ✅ CORREGIDO: UPDATE incluye piso y estado
                cursor.execute("""
                    UPDATE inventario 
                    SET usuario=?, oficina_id=?, pc_nombre=?, pc_usuario=?, 
                        windows=?, ip=?, marca=?, modelo=?, numero_serie=?,
                        mac_address=?, procesador=?, ram=?, disco=?,
                        motherboard=?, tarjeta_grafica=?, estado=?, piso=?, observaciones=?,
                        usa_ocs=?, contrasena=?
                    WHERE id=?
                """, (
                    data.get('usuario_persona', ''),    # ✅ CORRECCIÓN: usuario = Persona
                    oficina_id,
                    data.get('nombre_pc', ''),
                    data.get('pc_usuario', ''),         # ✅ CORRECCIÓN: pc_usuario = Usuario red
                    data.get('windows', ''),
                    data.get('ip_pc', ''),
                    data.get('marca', ''),
                    data.get('modelo', ''),
                    data.get('serie', ''),
                    data.get('mac_address', ''),
                    data.get('procesador', ''),
                    data.get('ram', ''),
                    data.get('disco', ''),
                    data.get('motherboard', ''),
                    data.get('tarjeta_grafica', ''),
                    data.get('estado', 'Ok'),           # ✅ CORREGIDO: Guarda estado
                    data.get('piso', ''),               # ✅ CORREGIDO: Guarda piso
                    data.get('observaciones', ''),
                    1 if data.get('usa_ocs') else 0,
                    data.get('contrasena', ''),
                    inventory_id
                ))
                conn.commit()
                print(f"✅ Registro {inventory_id} actualizado exitosamente")
                return cursor.rowcount > 0
        except Exception as e:
            print(f"❌ Error actualizando registro {inventory_id}: {e}")
            return False
    
    def add_inventory_record(self, data):
        """Agrega nuevo registro CON TODOS LOS CAMPOS incluyendo estado y piso"""
        try:
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                
                # Buscar oficina_id
                oficina_nombre = data.get('oficina', '')
                oficina_id = None
                if oficina_nombre:
                    try:
                        with self.get_oficinas_connection() as ofi_conn:
                            ofi_cursor = ofi_conn.cursor()
                            ofi_cursor.execute("SELECT id FROM oficinas WHERE nombre_oficina = ?", (oficina_nombre,))
                            result = ofi_cursor.fetchone()
                            if result:
                                oficina_id = result[0]
                    except Exception:
                        pass
                
                # ✅ CORREGIDO: INSERT incluye piso y estado
                cursor.execute("""
                    INSERT INTO inventario (
                        usuario, oficina_id, pc_nombre, pc_usuario, windows, ip, 
                        marca, modelo, numero_serie, mac_address, usa_ocs, contrasena,
                        procesador, ram, disco, motherboard, tarjeta_grafica, estado, piso, observaciones
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    data.get('usuario_persona', ''),    # ✅ CORRECCIÓN: usuario = Persona
                    oficina_id,
                    data.get('nombre_pc', ''),
                    data.get('pc_usuario', ''),         # ✅ CORRECCIÓN: pc_usuario = Usuario red
                    data.get('windows', ''),
                    data.get('ip_pc', ''),
                    data.get('marca', ''),
                    data.get('modelo', ''),
                    data.get('serie', ''),
                    data.get('mac_address', ''),
                    data.get('usa_ocs', 0),
                    data.get('contrasena', ''),
                    data.get('procesador', ''),
                    data.get('ram', ''),
                    data.get('disco', ''),
                    data.get('motherboard', ''),
                    data.get('tarjeta_grafica', ''),
                    data.get('estado', 'Ok'),           # ✅ CORREGIDO: Guarda estado
                    data.get('piso', ''),               # ✅ CORREGIDO: Guarda piso
                    data.get('observaciones', '')
                ))
                conn.commit()
                print(f"✅ Nuevo registro creado exitosamente")
                return cursor.lastrowid
        except Exception as e:
            print(f"❌ Error agregando registro: {e}")
            return None
    
    def delete_inventory_record(self, inventory_id):
        """Elimina un registro de la BD unificada"""
        try:
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM inventario WHERE id=?", (inventory_id,))
                conn.commit()
                return cursor.rowcount > 0
        except Exception as e:
            print(f"Error eliminando registro {inventory_id}: {e}")
            return False
    
    def get_duplicates_analysis(self):
        """Analiza duplicados considerando MAC Address también"""
        try:
            registros = self.get_all_inventory()
            duplicados = []
            
            series_vistas = {}
            ips_vistas = {}
            macs_vistas = {}
            nombres_vistos = {}
            
            for registro in registros:
                id_reg = registro['id']
                serie = registro['serie'].strip().lower() if registro['serie'] else ''
                ip = registro['ip_pc'].strip() if registro['ip_pc'] else ''
                mac = registro['mac_address'].strip().upper() if registro['mac_address'] else ''
                nombre_pc = registro['nombre_pc'].strip().lower() if registro['nombre_pc'] else ''
                
                # Analizar series
                if serie and len(serie) > 2:
                    if serie in series_vistas:
                        series_vistas[serie].append(id_reg)
                    else:
                        series_vistas[serie] = [id_reg]
                
                # Analizar IPs
                if ip and not ip.startswith(('127.', '169.254.')) and '.' in ip:
                    if ip in ips_vistas:
                        ips_vistas[ip].append(id_reg)
                    else:
                        ips_vistas[ip] = [id_reg]
                
                # Analizar MACs
                if mac and len(mac) > 10 and not mac.startswith('00:00:00'):
                    if mac in macs_vistas:
                        macs_vistas[mac].append(id_reg)
                    else:
                        macs_vistas[mac] = [id_reg]
                
                # Analizar nombres PC
                if nombre_pc and len(nombre_pc) > 2:
                    if nombre_pc in nombres_vistos:
                        nombres_vistos[nombre_pc].append(id_reg)
                    else:
                        nombres_vistos[nombre_pc] = [id_reg]
            
            ids_duplicados = set()
            
            # Marcar duplicados
            for ids in series_vistas.values():
                if len(ids) > 1:
                    ids_duplicados.update(ids)
            
            for ids in ips_vistas.values():
                if len(ids) > 1:
                    ids_duplicados.update(ids)
            
            for ids in macs_vistas.values():
                if len(ids) > 1:
                    ids_duplicados.update(ids)
            
            for ids in nombres_vistos.values():
                if len(ids) > 1:
                    ids_duplicados.update(ids)
            
            for registro in registros:
                if registro['id'] in ids_duplicados:
                    registro['is_duplicate'] = 1
                    duplicados.append(registro)
            
            return duplicados
            
        except Exception as e:
            print(f"Error analizando duplicados: {e}")
            return []
    
    def get_statistics(self):
        """Obtiene estadísticas del inventario INCLUYENDO estados"""
        try:
            stats = {}
            registros = self.get_all_inventory()
            
            stats['total_registros'] = len(registros)
            stats['duplicados'] = len(self.get_duplicates_analysis())
            
            # Estadísticas por oficina
            oficinas_count = {}
            for reg in registros:
                oficina = reg['oficina']
                oficinas_count[oficina] = oficinas_count.get(oficina, 0) + 1
            stats['por_oficina'] = dict(sorted(oficinas_count.items(), key=lambda x: x[1], reverse=True)[:10])
            
            # Estadísticas por marca
            marcas_count = {}
            for reg in registros:
                marca = reg['marca']
                if marca:
                    marcas_count[marca] = marcas_count.get(marca, 0) + 1
            stats['por_marca'] = dict(sorted(marcas_count.items(), key=lambda x: x[1], reverse=True))
            
            # Estadísticas por Windows (ACTUALIZADO)
            windows_count = {}
            for reg in registros:
                windows = reg.get('windows', '').strip() if reg.get('windows') else ''
                if windows and windows.lower() not in ['no especificado', 'n/a', '']:
                    windows_count[windows] = windows_count.get(windows, 0) + 1

            stats['por_windows'] = dict(sorted(windows_count.items(), key=lambda x: x[1], reverse=True))
            
            # Estadísticas de asignación - Asignadas vs Vacías (NUEVO)
            asignadas = 0
            vacias = 0
            for reg in registros:
                usuario = reg.get('usuario_persona', '').strip() if reg.get('usuario_persona') else ''
                if usuario and usuario.lower() not in ['', 'vacía', 'vacia', 'sin asignar']:
                    asignadas += 1
                else:
                    vacias += 1
            
            stats['asignadas'] = asignadas
            stats['vacias'] = vacias
            
            # Estadísticas por estado
            estados_count = {}
            for reg in registros:
                estado = reg['estado']
                estados_count[estado] = estados_count.get(estado, 0) + 1
            stats['por_estado'] = estados_count
            
            # Estadísticas específicas
            stats['desenchufadas'] = len([reg for reg in registros if reg['estado'] == 'Desenchufada'])
            stats['de_baja'] = len([reg for reg in registros if reg['estado'] == 'Baja'])
            stats['ok'] = len([reg for reg in registros if reg['estado'] == 'Ok'])
            
            # Estadísticas de MACs
            macs_registradas = len([reg for reg in registros if reg['mac_address']])
            stats['macs_registradas'] = macs_registradas
            stats['sin_mac'] = len(registros) - macs_registradas
            
            # Estadísticas de componentes
            stats['con_procesador'] = len([reg for reg in registros if reg['procesador']])
            stats['con_ram'] = len([reg for reg in registros if reg['ram']])
            stats['con_disco'] = len([reg for reg in registros if reg['disco']])
            
            stats['oficinas'] = len(set(reg['oficina'] for reg in registros if reg['oficina']))
            stats['marcas'] = len(marcas_count)
            
            return stats
            
        except Exception as e:
            print(f"Error obteniendo estadísticas: {e}")
            return {}
    # ========================= IMPRESORAS ==========================
    def get_marcas_impresoras(self):
        """Obtiene lista de marcas disponibles"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, nombre FROM marcas ORDER BY nombre")
                return [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]
        except Exception as e:
            print(f"Error obteniendo marcas: {e}")
            return []

    def get_modelos_por_marca(self, marca_id):
        """Obtiene modelos disponibles para una marca específica"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, nombre FROM modelos WHERE marca_id = ? ORDER BY nombre", (marca_id,))
                return [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]
        except Exception as e:
            print(f"Error obteniendo modelos: {e}")
            return []

    def get_toners_por_modelo(self, modelo_id):
        """Obtiene toners disponibles para un modelo específico"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, nombre FROM toners WHERE modelo_id = ? ORDER BY nombre", (modelo_id,))
                return [{'id': row[0], 'nombre': row[1]} for row in cursor.fetchall()]
        except Exception as e:
            print(f"Error obteniendo toners: {e}")
            return []

    def get_toner_por_marca_modelo(self, marca_id, modelo_id):
        """Obtiene toner específico para una combinación marca-modelo"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT nombre FROM toners 
                    WHERE marca_id = ? AND modelo_id = ?
                    LIMIT 1
                """, (marca_id, modelo_id))
                result = cursor.fetchone()
                return result[0] if result else None
        except Exception as e:
            print(f"Error obteniendo toner: {e}")
            return None
        
    def get_marca_nombre(self, marca_id):
        """Obtiene el nombre de una marca por su ID"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT nombre FROM marcas WHERE id = ?", (marca_id,))
                result = cursor.fetchone()
                return result[0] if result else ''
        except Exception as e:
            print(f"Error obteniendo nombre de marca: {e}")
            return ''

    def get_modelo_nombre(self, modelo_id):
        """Obtiene el nombre de un modelo por su ID"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT nombre FROM modelos WHERE id = ?", (modelo_id,))
                result = cursor.fetchone()
                return result[0] if result else ''
        except Exception as e:
            print(f"Error obteniendo nombre de modelo: {e}")
            return ''
    # ========================== NOTEBOOKS =======================================

    def _init_notebooks_db(self):
        """Inicializa la tabla de notebooks y su historial"""
        try:
            with self.get_notebooks_connection() as conn:
                cursor = conn.cursor()
                
                # Tabla principal de notebooks
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS notebooks (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        numero_serie TEXT UNIQUE NOT NULL,
                        marca TEXT,
                        modelo TEXT,
                        ip TEXT,
                        mac_address TEXT,
                        estado TEXT DEFAULT 'Guardada',
                        persona_actual TEXT,
                        fecha_entrega TIMESTAMP,
                        fecha_devolucion TIMESTAMP,
                        observaciones TEXT,
                        fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                
                # Tabla de historial de préstamos
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS notebooks_historial (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        notebook_id INTEGER NOT NULL,
                        numero_serie TEXT NOT NULL,
                        persona TEXT NOT NULL,
                        fecha_entrega TIMESTAMP NOT NULL,
                        fecha_devolucion TIMESTAMP,
                        observaciones TEXT,
                        FOREIGN KEY (notebook_id) REFERENCES notebooks(id)
                    )
                """)
                
                conn.commit()
                print("✅ Tablas de notebooks inicializadas correctamente")
                
        except Exception as e:
            print(f"❌ Error inicializando BD notebooks: {e}")

    def get_notebooks_connection(self):
        """Conexión a BD de notebooks"""
        try:
            conn = sqlite3.connect(self.notebooks_db_path, timeout=5)
            conn.execute("SELECT 1")
            return conn
        except Exception:
            if "16.1.1.118" in self.notebooks_db_path:
                self.notebooks_db_path = app.config['LOCAL_NOTEBOOKS_DB']
                self._init_notebooks_db()
            return sqlite3.connect(self.notebooks_db_path)

    def trasladar_a_notebooks(self, numero_serie):
        """Traslada una notebook del inventario general a la tabla de notebooks"""
        try:
            # Buscar en inventario por número de serie
            with self.get_inventario_connection() as inv_conn:
                inv_cursor = inv_conn.cursor()
                inv_cursor.execute("""
                    SELECT id, marca, modelo, ip, mac_address, numero_serie, pc_nombre
                    FROM inventario
                    WHERE numero_serie = ?
                """, (numero_serie,))
                
                registro = inv_cursor.fetchone()
                
                if not registro:
                    return False, "Número de serie no encontrado en inventario"
                
                id_inv, marca, modelo, ip, mac, serie, pc_nombre = registro
            
            # Insertar en notebooks
            with self.get_notebooks_connection() as nb_conn:
                nb_cursor = nb_conn.cursor()
                
                # Verificar si ya existe
                nb_cursor.execute("SELECT id FROM notebooks WHERE numero_serie = ?", (numero_serie,))
                if nb_cursor.fetchone():
                    return False, "Esta notebook ya está en el sistema de notebooks"
                
                nb_cursor.execute("""
                    INSERT INTO notebooks (numero_serie, marca, modelo, ip, mac_address, pc_nombre, estado)
                    VALUES (?, ?, ?, ?, ?, ?, 'Guardada')
                """, (serie, marca, modelo, ip, mac, pc_nombre))
                
                nb_conn.commit()
                notebook_id = nb_cursor.lastrowid
            
            # Eliminar del inventario
            with self.get_inventario_connection() as inv_conn:
                inv_cursor = inv_conn.cursor()
                inv_cursor.execute("DELETE FROM inventario WHERE id = ?", (id_inv,))
                inv_conn.commit()
            
            print(f"✅ Notebook {numero_serie} trasladada correctamente")
            return True, notebook_id
            
        except Exception as e:
            print(f"Error trasladando notebook: {e}")
            return False, str(e)

    def get_all_notebooks(self):
        """Obtiene todas las notebooks"""
        try:
            with self.get_notebooks_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, numero_serie, marca, modelo, ip, mac_address, 
                        estado, persona_actual, fecha_entrega, fecha_devolucion,
                        observaciones, fecha_registro, pc_nombre
                    FROM notebooks
                    ORDER BY estado DESC, fecha_registro DESC
                """)
                
                notebooks = []
                for row in cursor.fetchall():
                    notebooks.append({
                        'id': row[0],
                        'numero_serie': row[1] or '',
                        'marca': row[2] or '',
                        'modelo': row[3] or '',
                        'ip': row[4] or '',
                        'mac_address': row[5] or '',
                        'estado': row[6] or 'Guardada',
                        'persona_actual': row[7] or '',
                        'fecha_entrega': row[8] or '',
                        'fecha_devolucion': row[9] or '',
                        'observaciones': row[10] or '',
                        'fecha_registro': row[11] or '',
                        'pc_nombre': row[12] or ''
                    })
                
                return notebooks
                
        except Exception as e:
            print(f"Error obteniendo notebooks: {e}")
            return []

    def get_notebook_by_id(self, notebook_id):
        """Obtiene una notebook específica"""
        try:
            with self.get_notebooks_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, numero_serie, marca, modelo, ip, mac_address, 
                        estado, persona_actual, fecha_entrega, fecha_devolucion,
                        observaciones, fecha_registro, pc_nombre
                    FROM notebooks
                    WHERE id = ?
                """, (notebook_id,))
                
                row = cursor.fetchone()
                if row:
                    return {
                        'id': row[0],
                        'numero_serie': row[1] or '',
                        'marca': row[2] or '',
                        'modelo': row[3] or '',
                        'ip': row[4] or '',
                        'mac_address': row[5] or '',
                        'estado': row[6] or 'Guardada',
                        'persona_actual': row[7] or '',
                        'fecha_entrega': row[8] or '',
                        'fecha_devolucion': row[9] or '',
                        'observaciones': row[10] or '',
                        'fecha_registro': row[11] or '',
                        'pc_nombre': row[12] or ''
                    }
                return None
                
        except Exception as e:
            print(f"Error obteniendo notebook {notebook_id}: {e}")
            return None

    def prestar_notebook(self, notebook_id, persona, observaciones=''):
        """Registra el préstamo de una notebook"""
        try:
            fecha_entrega = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            with self.get_notebooks_connection() as conn:
                cursor = conn.cursor()
                
                # Actualizar notebook
                cursor.execute("""
                    UPDATE notebooks
                    SET estado = 'Prestada',
                        persona_actual = ?,
                        fecha_entrega = ?,
                        fecha_devolucion = NULL,
                        observaciones = ?
                    WHERE id = ?
                """, (persona, fecha_entrega, observaciones, notebook_id))
                
                # Obtener número de serie
                cursor.execute("SELECT numero_serie FROM notebooks WHERE id = ?", (notebook_id,))
                numero_serie = cursor.fetchone()[0]
                
                # Registrar en historial
                cursor.execute("""
                    INSERT INTO notebooks_historial 
                    (notebook_id, numero_serie, persona, fecha_entrega, observaciones)
                    VALUES (?, ?, ?, ?, ?)
                """, (notebook_id, numero_serie, persona, fecha_entrega, observaciones))
                
                conn.commit()
                return True
                
        except Exception as e:
            print(f"Error prestando notebook: {e}")
            return False

    def devolver_notebook(self, notebook_id, observaciones=''):
        """Registra la devolución de una notebook"""
        try:
            fecha_devolucion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            with self.get_notebooks_connection() as conn:
                cursor = conn.cursor()
                
                # Actualizar último registro del historial
                cursor.execute("""
                    UPDATE notebooks_historial
                    SET fecha_devolucion = ?,
                        observaciones = observaciones || '' || ?
                    WHERE notebook_id = ? AND fecha_devolucion IS NULL
                """, (fecha_devolucion, observaciones, notebook_id))
                
                # Actualizar notebook
                cursor.execute("""
                    UPDATE notebooks
                    SET estado = 'Guardada',
                        persona_actual = NULL,
                        fecha_devolucion = ?,
                        observaciones = ?
                    WHERE id = ?
                """, (fecha_devolucion, observaciones, notebook_id))
                
                conn.commit()
                return True
                
        except Exception as e:
            print(f"Error devolviendo notebook: {e}")
            return False

    def get_notebook_historial(self, notebook_id):
        """Obtiene el historial de préstamos de una notebook"""
        try:
            with self.get_notebooks_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, persona, fecha_entrega, fecha_devolucion, observaciones
                    FROM notebooks_historial
                    WHERE notebook_id = ?
                    ORDER BY fecha_entrega DESC
                """, (notebook_id,))
                
                historial = []
                for row in cursor.fetchall():
                    historial.append({
                        'id': row[0],
                        'persona': row[1],
                        'fecha_entrega': row[2],
                        'fecha_devolucion': row[3],
                        'observaciones': row[4] or ''
                    })
                
                return historial
                
        except Exception as e:
            print(f"Error obteniendo historial: {e}")
            return []

    def actualizar_notebook(self, notebook_id, numero_serie, pc_nombre, marca, modelo, ip, mac_address, observaciones, estado=None):
        """Actualiza la información de una notebook"""
        try:
            with self.get_notebooks_connection() as conn:
                cursor = conn.cursor()
                
                # Verificar si el número de serie ya existe en otra notebook
                cursor.execute("""
                    SELECT id FROM notebooks 
                    WHERE numero_serie = ? AND id != ?
                """, (numero_serie, notebook_id))
                
                if cursor.fetchone():
                    print(f"Error: El número de serie {numero_serie} ya existe en otra notebook")
                    return False
                
                # Si se proporciona estado, actualizarlo también
                if estado:
                    # Si el estado es "En uso interno", asegurarse de que persona_actual sea "Computos"
                    if estado == 'En uso interno':
                        cursor.execute("""
                            UPDATE notebooks
                            SET numero_serie = ?,
                                pc_nombre = ?,
                                marca = ?,
                                modelo = ?,
                                ip = ?,
                                mac_address = ?,
                                observaciones = ?,
                                estado = ?,
                                persona_actual = 'Computos'
                            WHERE id = ?
                        """, (numero_serie, pc_nombre, marca, modelo, ip, mac_address, observaciones, estado, notebook_id))
                    
                    # Si el estado es "Prestada", mantener la persona actual
                    elif estado == 'Prestada':
                        cursor.execute("""
                            UPDATE notebooks
                            SET numero_serie = ?,
                                pc_nombre = ?,
                                marca = ?,
                                modelo = ?,
                                ip = ?,
                                mac_address = ?,
                                observaciones = ?,
                                estado = ?
                            WHERE id = ?
                        """, (numero_serie, pc_nombre, marca, modelo, ip, mac_address, observaciones, estado, notebook_id))
                    
                    # Para otros estados, limpiar persona_actual
                    else:
                        cursor.execute("""
                            UPDATE notebooks
                            SET numero_serie = ?,
                                pc_nombre = ?,
                                marca = ?,
                                modelo = ?,
                                ip = ?,
                                mac_address = ?,
                                observaciones = ?,
                                estado = ?,
                                persona_actual = NULL
                            WHERE id = ?
                        """, (numero_serie, pc_nombre, marca, modelo, ip, mac_address, observaciones, estado, notebook_id))
                else:
                    # Actualizar sin cambiar el estado
                    cursor.execute("""
                        UPDATE notebooks
                        SET numero_serie = ?,
                            pc_nombre = ?,
                            marca = ?,
                            modelo = ?,
                            ip = ?,
                            mac_address = ?,
                            observaciones = ?
                        WHERE id = ?
                    """, (numero_serie, pc_nombre, marca, modelo, ip, mac_address, observaciones, notebook_id))
                
                conn.commit()
                
                if cursor.rowcount > 0:
                    print(f"✅ Notebook {notebook_id} actualizada correctamente")
                    return True
                else:
                    print(f"⚠️ No se encontró la notebook {notebook_id}")
                    return False
                    
        except Exception as e:
            print(f"❌ Error actualizando notebook: {e}")
            return False
        
    def eliminar_notebook(self, notebook_id):
        """Elimina una notebook y su historial"""
        try:
            with self.get_notebooks_connection() as conn:
                cursor = conn.cursor()
                
                # Obtener información de la notebook antes de borrar
                cursor.execute("SELECT numero_serie, estado FROM notebooks WHERE id = ?", (notebook_id,))
                result = cursor.fetchone()
                
                if not result:
                    print(f"⚠️ Notebook {notebook_id} no encontrada")
                    return False
                
                numero_serie, estado = result
                
                # Eliminar historial primero
                cursor.execute("DELETE FROM notebooks_historial WHERE notebook_id = ?", (notebook_id,))
                historial_borrados = cursor.rowcount
                
                # Eliminar notebook
                cursor.execute("DELETE FROM notebooks WHERE id = ?", (notebook_id,))
                
                conn.commit()
                
                if cursor.rowcount > 0:
                    print(f"✅ Notebook {numero_serie} eliminada correctamente")
                    print(f"   - Registros de historial eliminados: {historial_borrados}")
                    return True
                else:
                    print(f"⚠️ No se pudo eliminar la notebook {notebook_id}")
                    return False
                    
        except Exception as e:
            print(f"❌ Error eliminando notebook: {e}")
            return False

    def cambiar_estado_notebook(self, notebook_id, nuevo_estado, observaciones='', persona=None):
        """Cambia el estado de una notebook"""
        try:
            with self.get_notebooks_connection() as conn:
                cursor = conn.cursor()
                
                # Determinar qué hacer con persona_actual según el nuevo estado
                if nuevo_estado == 'Prestada':
                    # Si es Prestada, debe tener una persona asignada
                    if not persona or persona.strip() == '':
                        print(f"⚠️ Estado 'Prestada' requiere una persona asignada")
                        return False
                    cursor.execute("""
                        UPDATE notebooks
                        SET estado = ?,
                            persona_actual = ?,
                            observaciones = ?,
                            fecha_entrega = datetime('now', 'localtime')
                        WHERE id = ?
                    """, (nuevo_estado, persona.strip(), observaciones, notebook_id))
                    
                elif nuevo_estado == 'En uso interno':
                    # Si es En uso interno, asignar automáticamente a "Computos"
                    # Usar el valor de persona si viene del formulario, sino "Computos"
                    persona_asignar = persona.strip() if persona and persona.strip() else 'Computos'
                    cursor.execute("""
                        UPDATE notebooks
                        SET estado = ?,
                            persona_actual = ?,
                            observaciones = ?,
                            fecha_entrega = datetime('now', 'localtime')
                        WHERE id = ?
                    """, (nuevo_estado, persona_asignar, observaciones, notebook_id))
                    
                elif nuevo_estado == 'Guardada':
                    # Estado Guardada: limpiar persona_actual y fecha
                    cursor.execute("""
                        UPDATE notebooks
                        SET estado = ?,
                            persona_actual = NULL,
                            observaciones = ?,
                            fecha_entrega = NULL
                        WHERE id = ?
                    """, (nuevo_estado, observaciones, notebook_id))
                    
                else:
                    # Para cualquier otro estado, limpiar persona_actual
                    cursor.execute("""
                        UPDATE notebooks
                        SET estado = ?,
                            persona_actual = NULL,
                            observaciones = ?,
                            fecha_entrega = NULL
                        WHERE id = ?
                    """, (nuevo_estado, observaciones, notebook_id))
                
                conn.commit()
                
                if cursor.rowcount > 0:
                    print(f"✅ Notebook {notebook_id} cambiada a estado: {nuevo_estado}")
                    if nuevo_estado == 'En uso interno':
                        print(f"   - Asignada a: {persona_asignar if 'persona_asignar' in locals() else 'Computos'}")
                    elif nuevo_estado == 'Prestada' and persona:
                        print(f"   - Asignada a: {persona}")
                    return True
                else:
                    print(f"⚠️ No se encontró la notebook {notebook_id}")
                    return False
                    
        except Exception as e:
            print(f"❌ Error cambiando estado: {e}")
            return False
    
    # ==================== MÉTODOS PARA IMPRESORAS ====================
    
    def _init_impresoras_db(self):
        """Inicializa la tabla de impresoras"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS impresoras (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        oficina_id INTEGER,
                        oficina_nombre TEXT,
                        piso TEXT,
                        marca TEXT NOT NULL,
                        modelo TEXT,
                        numero_serie TEXT,
                        tipo_conexion TEXT NOT NULL DEFAULT 'red',
                        ip TEXT,
                        toner TEXT,
                        observaciones TEXT,
                        fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        usuario_registro TEXT,
                        estado TEXT DEFAULT 'Activa'
                    )
                """)
                conn.commit()
                print("✓ Tabla impresoras inicializada correctamente")
        except Exception as e:
            print(f"Error inicializando BD de impresoras: {e}")

    def get_all_printers(self):
        """Obtiene todas las impresoras del inventario"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, oficina_id, oficina_nombre, piso, marca, modelo, numero_serie,
                        tipo_conexion, ip, toner, observaciones,
                        fecha_registro, usuario_registro, estado
                    FROM impresoras
                    ORDER BY oficina_nombre, piso, marca
                """)
                
                impresoras_raw = cursor.fetchall()
                impresoras = []
                
                for row in impresoras_raw:
                    (id_imp, oficina_id, oficina_nombre, piso, marca, modelo, numero_serie,
                    tipo_conexion, ip, toner, observaciones,
                    fecha_registro, usuario_registro, estado) = row
                    
                    impresoras.append({
                        'id': id_imp,
                        'oficina_id': oficina_id,
                        'oficina': oficina_nombre or '',
                        'piso': str(piso) if piso else '',
                        'marca': marca or '',
                        'modelo': modelo or '',
                        'numero_serie': numero_serie or '',
                        'tipo_conexion': tipo_conexion or 'red',
                        'ip': ip or '',
                        'toner': toner or '',
                        'observaciones': observaciones or '',
                        'fecha_registro': fecha_registro or '',
                        'usuario_registro': usuario_registro or '',
                        'estado': estado or 'Activa'
                    })
                
                return impresoras
                
        except Exception as e:
            print(f"Error obteniendo impresoras: {e}")
            return []

    def get_printer_by_id(self, printer_id):
        """Obtiene una impresora específica"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT id, oficina_id, oficina_nombre, piso, marca, modelo, numero_serie,
                        tipo_conexion, ip, toner, observaciones,
                        fecha_registro, usuario_registro, estado
                    FROM impresoras
                    WHERE id = ?
                """, (printer_id,))
                
                row = cursor.fetchone()
                if row:
                    tipo_conexion = row[7] or 'red'
                    ip = row[8] or ''
                    
                    return {
                        'id': row[0],
                        'oficina_id': row[1],
                        'oficina': row[2] or '',
                        'piso': str(row[3]) if row[3] else '',
                        'marca': row[4] or '',
                        'modelo': row[5] or '',
                        'numero_serie': row[6] or '',
                        'tipo_conexion': tipo_conexion,
                        'ip': ip,
                        'toner': row[9] or '',
                        'observaciones': row[10] or '',
                        'fecha_registro': row[11] or '',
                        'usuario_registro': row[12] or '',
                        'estado': row[13] or 'Activa',
                        # Variables calculadas para los templates
                        'tiene_ip': tipo_conexion == 'red',
                        'ip_impresora': ip if tipo_conexion == 'red' else '',
                        'ip_pc_madre': ip if tipo_conexion == 'compartida' else ''
                    }
                return None
                
        except Exception as e:
            print(f"Error obteniendo impresora {printer_id}: {e}")
            return None

    def add_printer(self, data):
        """Agrega nueva impresora al inventario"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO impresoras (
                        oficina_id, oficina_nombre, piso, marca, modelo, numero_serie,
                        tipo_conexion, ip, toner, observaciones,
                        usuario_registro, estado
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    data.get('oficina_id'),
                    data.get('oficina'),
                    data.get('piso'),
                    data.get('marca'),
                    data.get('modelo', ''),
                    data.get('numero_serie', ''),
                    data.get('tipo_conexion', 'red'),
                    data.get('ip', ''),
                    data.get('toner', ''),
                    data.get('observaciones', ''),
                    data.get('usuario_registro', 'Web'),
                    data.get('estado', 'Activa')
                ))
                conn.commit()
                print(f"Impresora agregada exitosamente: {data.get('marca')} {data.get('modelo')}")
                return cursor.lastrowid
        except Exception as e:
            print(f"Error agregando impresora: {e}")
            return None

    def update_printer(self, printer_id, data):
        """Actualiza una impresora existente"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE impresoras 
                    SET oficina_id=?, oficina_nombre=?, piso=?, marca=?, modelo=?, numero_serie=?,
                        tipo_conexion=?, ip=?, toner=?, observaciones=?, estado=?
                    WHERE id=?
                """, (
                    data.get('oficina_id'),
                    data.get('oficina'),
                    data.get('piso'),
                    data.get('marca'),
                    data.get('modelo', ''),
                    data.get('numero_serie', ''),
                    data.get('tipo_conexion', 'red'),
                    data.get('ip', ''),
                    data.get('toner', ''),
                    data.get('observaciones', ''),
                    data.get('estado', 'Activa'),
                    printer_id
                ))
                conn.commit()
                print(f"Impresora {printer_id} actualizada exitosamente")
                return cursor.rowcount > 0
        except Exception as e:
            print(f"Error actualizando impresora {printer_id}: {e}")
            return False

    def delete_printer(self, printer_id):
        """Elimina una impresora del inventario"""
        try:
            with self.get_impresoras_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM impresoras WHERE id=?", (printer_id,))
                conn.commit()
                print(f"Impresora {printer_id} eliminada exitosamente")
                return cursor.rowcount > 0
        except Exception as e:
            print(f"Error eliminando impresora {printer_id}: {e}")
            return False

    def get_printers_statistics(self):
        """Obtiene estadísticas de las impresoras"""
        try:
            impresoras = self.get_all_printers()
            stats = {}
            
            stats['total_impresoras'] = len(impresoras)
            
            # Por oficina
            oficinas_count = {}
            for imp in impresoras:
                oficina = imp['oficina']
                if oficina:
                    oficinas_count[oficina] = oficinas_count.get(oficina, 0) + 1
            stats['por_oficina'] = dict(sorted(oficinas_count.items(), key=lambda x: x[1], reverse=True)[:10])
            
            # Por marca
            marcas_count = {}
            for imp in impresoras:
                marca = imp['marca']
                if marca:
                    marcas_count[marca] = marcas_count.get(marca, 0) + 1
            stats['por_marca'] = dict(sorted(marcas_count.items(), key=lambda x: x[1], reverse=True))
            
            # Por estado
            estados_count = {}
            for imp in impresoras:
                estado = imp['estado']
                estados_count[estado] = estados_count.get(estado, 0) + 1
            stats['por_estado'] = estados_count
            
            # Por tipo de conexión
            conexion_count = {}
            for imp in impresoras:
                tipo = imp['tipo_conexion']
                if tipo == 'red':
                    conexion_count['Con IP propia'] = conexion_count.get('Con IP propia', 0) + 1
                else:
                    conexion_count['Compartida'] = conexion_count.get('Compartida', 0) + 1
            stats['conectividad'] = conexion_count
            
            return stats
            
        except Exception as e:
            print(f"Error obteniendo estadísticas de impresoras: {e}")
            return {}

    def agregar_columna_pc_nombre(self):
            """Agrega la columna pc_nombre a la tabla notebooks si no existe"""
            try:
                with self.get_notebooks_connection() as conn:
                    cursor = conn.cursor()
                    
                    # Verificar si la columna ya existe
                    cursor.execute("PRAGMA table_info(notebooks)")
                    columnas = [info[1] for info in cursor.fetchall()]
                    
                    if 'pc_nombre' not in columnas:
                        cursor.execute("ALTER TABLE notebooks ADD COLUMN pc_nombre TEXT")
                        conn.commit()
                        print("✅ Columna pc_nombre agregada exitosamente")
                    else:
                        print("ℹ️ La columna pc_nombre ya existe")
                        
            except Exception as e:
                print(f"❌ Error agregando columna pc_nombre: {e}")

    # ==================== MÉTODOS COMPARTIDOS ====================
    
    def _get_oficina_name_from_cne(self, oficina_id):
        """Obtiene nombre de oficina desde BD CNE por ID"""
        try:
            with self.get_oficinas_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT nombre_oficina FROM oficinas WHERE id = ?", (oficina_id,))
                result = cursor.fetchone()
                return result[0] if result else None
        except:
            return None
    
    def get_oficinas_list(self):
        """Obtiene lista de oficinas desde BD CNE"""
        try:
            with self.get_oficinas_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT nombre_oficina FROM oficinas ORDER BY nombre_oficina")
                return [row[0] for row in cursor.fetchall()]
        except Exception:
            return []
    
    def get_oficina_with_piso(self, oficina_nombre):
        """Obtiene oficina con su piso desde BD CNE"""
        try:
            with self.get_oficinas_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, piso FROM oficinas WHERE nombre_oficina = ?", (oficina_nombre,))
                result = cursor.fetchone()
                if result:
                    return {'id': result[0], 'piso': result[1]}
                return None
        except Exception:
            return None

    def buscar_por_mac(self, mac_address):
        """Busca registros por dirección MAC"""
        try:
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT inv.id, inv.usuario, inv.oficina_id, inv.pc_nombre, inv.mac_address,
                           ofi.nombre as oficina_nombre 
                    FROM inventario inv
                    LEFT JOIN oficinas ofi ON inv.oficina_id = ofi.id
                    WHERE inv.mac_address = ?
                """, (mac_address,))
                return cursor.fetchall()
        except Exception as e:
            print(f"Error buscando por MAC {mac_address}: {e}")
            return []

    def get_macs_registradas(self):
        """Obtiene lista de todas las MACs registradas"""
        try:
            with self.get_inventario_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT DISTINCT mac_address 
                    FROM inventario 
                    WHERE mac_address IS NOT NULL AND mac_address != ''
                    ORDER BY mac_address
                """)
                return [row[0] for row in cursor.fetchall()]
        except Exception as e:
            print(f"Error obteniendo MACs registradas: {e}")
            return []

# Instancia global
unified_web_db = UnifiedWebDatabaseManager()

# ==================== RUTAS PRINCIPALES ====================

@app.route('/status')
def status():
    """Endpoint para verificar el estado de todas las bases de datos"""
    try:
        # Contar registros usando los métodos correctos
        inventario_list = unified_web_db.get_all_inventory()
        oficinas_list = unified_web_db.get_oficinas_list()
        impresoras_list = unified_web_db.get_all_printers()
        
        inventario_count = len(inventario_list) if inventario_list else 0
        oficinas_count = len(oficinas_list) if oficinas_list else 0
        impresoras_count = len(impresoras_list) if impresoras_list else 0
        
        # Contadores específicos
        macs_registradas = len([reg for reg in inventario_list if reg.get('mac_address')])
        desenchufadas = len([reg for reg in inventario_list if reg.get('estado') == 'Desenchufada'])
        de_baja = len([reg for reg in inventario_list if reg.get('estado') == 'Baja'])
        
        oficinas_status = "OK" if oficinas_count else "No disponible"
        impresoras_status = "OK" if impresoras_count else "No disponible"
        
        # Determinar tipo de conexión
        status_text = "Red" if getattr(unified_web_db, "is_shared", False) else "Local"
        database_type = "shared" if getattr(unified_web_db, "is_shared", False) else "local"
        
        return jsonify({
            "estado": "Conectado",
            "database_type": database_type,
            "inventario_path": getattr(unified_web_db, "inventario_db_path", "Desconocido"),
            "oficinas_path": getattr(unified_web_db, "oficinas_db_path", "Desconocido"),
            "impresoras_path": getattr(unified_web_db, "impresoras_db_path", "Desconocido"),
            "registros_inventario": inventario_count,
            "macs_registradas": macs_registradas,
            "desenchufadas": desenchufadas,
            "de_baja": de_baja,
            "oficinas_cne": oficinas_count,
            "impresoras_inventario": impresoras_count,
            "oficinas_status": oficinas_status,
            "impresoras_status": impresoras_status,
            "status_text": status_text
        })
    except Exception as e:
        return jsonify({
            "estado": "Error",
            "database_type": "error",
            "error": str(e)
        })

@app.route('/')
def index():
    try:
        # Obtener parámetros de ordenamiento
        sort_by = request.args.get('sort', 'default')  # 'default' para orden personalizado
        sort_order = request.args.get('order', 'asc')
        
        registros = unified_web_db.get_all_inventory()
        
        # Ordenar registros
        if registros:
            reverse = (sort_order == 'desc')
            
            if sort_by == 'default':
                # Ordenamiento por defecto: Piso -> Oficina -> IP
                def default_sort_key(registro):
                    # Convertir piso a número, si falla usar 999 para ponerlo al final
                    try:
                        piso = int(registro.get('piso', 999) or 999)
                    except (ValueError, TypeError):
                        piso = 999
                    
                    oficina = (registro.get('oficina') or '').lower()
                    
                    # Convertir IP a tupla de números para ordenamiento correcto
                    ip = registro.get('ip_pc') or ''
                    try:
                        ip_parts = tuple(int(x) for x in ip.split('.'))
                    except:
                        ip_parts = (999, 999, 999, 999)
                    
                    return (piso, oficina, ip_parts)
                
                registros = sorted(registros, key=default_sort_key, reverse=reverse)
            else:
                # Ordenamiento por columna específica
                def safe_sort_key(registro):
                    value = registro.get(sort_by, '')
                    return (value or '').lower() if isinstance(value, str) else (value or 0)
                
                try:
                    registros = sorted(registros, key=safe_sort_key, reverse=reverse)
                except Exception as e:
                    print(f"Error ordenando por {sort_by}: {e}")
        
        stats = unified_web_db.get_statistics()
        
        return render_template('dashboard.html', 
                             registros=registros, 
                             stats=stats,
                             current_sort=sort_by,
                             current_order=sort_order)
    except Exception as e:
        flash(f'Error cargando datos: {e}', 'error')
        return render_template('dashboard.html', registros=[], stats={})

@app.route('/duplicados')
def ver_duplicados():
    try:
        duplicados = unified_web_db.get_duplicates_analysis()
        return render_template('duplicados.html', duplicados=duplicados)
    except Exception as e:
        flash(f'Error analizando duplicados: {e}', 'error')
        return render_template('duplicados.html', duplicados=[])

@app.route('/estadisticas')
def ver_estadisticas():
    try:
        stats = unified_web_db.get_statistics()
        
        # DEBUG TEMPORAL
        print("=" * 50)
        print("DEBUG ESTADÍSTICAS:")
        print(f"Total registros: {stats.get('total_registros')}")
        print(f"Por Windows: {stats.get('por_windows')}")
        print(f"Asignadas: {stats.get('asignadas')}")
        print(f"Vacías: {stats.get('vacias')}")
        print("=" * 50)
        
        return render_template('estadisticas.html', stats=stats)
    except Exception as e:
        flash(f'Error cargando estadísticas: {e}', 'error')
        return render_template('estadisticas.html', stats={})

@app.route('/registro/<int:registro_id>')
def ver_registro(registro_id):
    try:
        registro = unified_web_db.get_inventory_by_id(registro_id)
        if not registro:
            flash('Registro no encontrado', 'error')
            return redirect(url_for('index'))
        return render_template('registro_detalle.html', registro=registro)
    except Exception as e:
        flash(f'Error cargando registro: {e}', 'error')
        return redirect(url_for('index'))

@app.route('/registro/<int:registro_id>/editar', methods=['GET','POST'])
def editar_registro(registro_id):
    if request.method == 'GET':
        try:
            registro = unified_web_db.get_inventory_by_id(registro_id)
            if not registro:
                flash('Registro no encontrado', 'error')
                return redirect(url_for('index'))
            oficinas = unified_web_db.get_oficinas_list()
            return render_template('registro_editar.html', registro=registro, oficinas=oficinas)
        except Exception as e:
            flash(f'Error cargando registro: {e}', 'error')
            return redirect(url_for('index'))
    else:
        try:
            data = {
                'usuario_persona': request.form.get('usuario_persona', ''),    # ✅ CORRECCIÓN: Persona
                'oficina': request.form.get('oficina', ''),
                'piso': request.form.get('piso', ''),
                'pc_usuario': request.form.get('pc_usuario', ''),             # ✅ CORRECCIÓN: Usuario Red
                'nombre_pc': request.form.get('nombre_pc', ''),
                'ip_pc': request.form.get('ip_pc', ''),
                'mac_address': request.form.get('mac_address', ''),
                'marca': request.form.get('marca', ''),
                'modelo': request.form.get('modelo', ''),
                'serie': request.form.get('serie', ''),
                'procesador': request.form.get('procesador', ''),
                'ram': request.form.get('ram', ''),
                'disco': request.form.get('disco', ''),
                'motherboard': request.form.get('motherboard', ''),
                'tarjeta_grafica': request.form.get('tarjeta_grafica', ''),
                'windows': request.form.get('windows', ''),
                'estado': request.form.get('estado', 'Ok'),
                'usa_ocs': request.form.get('usa_ocs'),
                'contrasena': request.form.get('contrasena', ''),
                'observaciones': request.form.get('observaciones', '')
            }
            success = unified_web_db.update_inventory_record(registro_id, data)
            if success:
                flash('Registro actualizado exitosamente', 'success')
            else:
                flash('Error actualizando registro', 'error')
        except Exception as e:
            flash(f'Error actualizando registro: {e}', 'error')
        return redirect(url_for('ver_registro', registro_id=registro_id))

@app.route('/registro/nuevo', methods=['GET', 'POST'])
def nuevo_registro():
    """Formulario para agregar nuevo registro de inventario"""
    if request.method == 'GET':
        oficinas = unified_web_db.get_oficinas_list()
        return render_template('registro_nuevo.html', oficinas=oficinas)
    else:
        try:
            data = {
                'usuario_persona': request.form.get('usuario_persona', ''),    # ✅ CORRECCIÓN: Persona
                'oficina': request.form.get('oficina', ''),
                'pc_usuario': request.form.get('pc_usuario', ''),             # ✅ CORRECCIÓN: Usuario Red
                'nombre_pc': request.form.get('nombre_pc', ''),
                'ip_pc': request.form.get('ip_pc', ''),
                'mac_address': request.form.get('mac_address', ''),
                'marca': request.form.get('marca', ''),
                'modelo': request.form.get('modelo', ''),
                'serie': request.form.get('serie', ''),
                'procesador': request.form.get('procesador', ''),
                'ram': request.form.get('ram', ''),
                'disco': request.form.get('disco', ''),
                'motherboard': request.form.get('motherboard', ''),
                'tarjeta_grafica': request.form.get('tarjeta_grafica', ''),
                'windows': request.form.get('windows', ''),
                'estado': request.form.get('estado', 'Ok'),
                'usa_ocs': 1 if request.form.get('usa_ocs') else 0,
                'contrasena': request.form.get('contrasena', ''),
                'observaciones': request.form.get('observaciones', '')
            }
            
            record_id = unified_web_db.add_inventory_record(data)
            if record_id:
                flash('Registro agregado exitosamente', 'success')
                return redirect(url_for('ver_registro', registro_id=record_id))
            else:
                flash('Error agregando registro', 'error')
                
        except Exception as e:
            flash(f'Error agregando registro: {e}', 'error')
        
        # Si hay error, volver al formulario
        oficinas = unified_web_db.get_oficinas_list()
        return render_template('registro_nuevo.html', oficinas=oficinas)

@app.route('/registro/<int:registro_id>/eliminar', methods=['POST'])
def eliminar_registro(registro_id):
    try:
        success = unified_web_db.delete_inventory_record(registro_id)
        if success:
            flash('Registro eliminado exitosamente', 'success')
        else:
            flash('Error eliminando registro', 'error')
    except Exception as e:
        flash(f'Error eliminando registro: {e}', 'error')
    return redirect(url_for('index'))

@app.route('/forzar_duplicados')
def forzar_duplicados():
    """Fuerza el análisis de duplicados"""
    try:
        duplicados = unified_web_db.get_duplicates_analysis()
        flash(f'Análisis de duplicados completado: {len(duplicados)} duplicados encontrados', 'success')
    except Exception as e:
        flash(f'Error en análisis de duplicados: {e}', 'error')
    return redirect(url_for('ver_duplicados'))

@app.route('/exportar')
def exportar_excel():
    try:
        registros = unified_web_db.get_all_inventory()
        
        # Crear workbook de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        # Definir headers
        headers = [
            'ID', 'Fecha', 'Persona (Usuario)', 'Oficina', 'Piso', 'Usuario Red', 
            'Nombre PC', 'IP', 'MAC Address', 'Marca', 'Modelo', 'Serie', 
            'Procesador', 'RAM', 'Disco',
            'Windows', 'Estado', 'Usa OCS', 'Observaciones'
        ]
        
        # Estilos para el header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir headers con estilo
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_num, r in enumerate(registros, 2):
            ws.cell(row=row_num, column=1, value=r.get('id', ''))
            ws.cell(row=row_num, column=2, value=r.get('fecha_creacion', '')[:19] if r.get('fecha_creacion') else '')
            ws.cell(row=row_num, column=3, value=r.get('usuario_persona', ''))
            ws.cell(row=row_num, column=4, value=r.get('oficina', ''))
            ws.cell(row=row_num, column=5, value=r.get('piso', ''))
            ws.cell(row=row_num, column=6, value=r.get('pc_usuario', ''))
            ws.cell(row=row_num, column=7, value=r.get('nombre_pc', ''))
            ws.cell(row=row_num, column=8, value=r.get('ip_pc', ''))
            ws.cell(row=row_num, column=9, value=r.get('mac_address', ''))
            ws.cell(row=row_num, column=10, value=r.get('marca', ''))
            ws.cell(row=row_num, column=11, value=r.get('modelo', ''))
            ws.cell(row=row_num, column=12, value=r.get('serie', ''))
            ws.cell(row=row_num, column=13, value=r.get('procesador', ''))
            ws.cell(row=row_num, column=14, value=r.get('ram', ''))
            ws.cell(row=row_num, column=15, value=r.get('disco', ''))
            ws.cell(row=row_num, column=16, value=r.get('windows', ''))
            ws.cell(row=row_num, column=17, value=r.get('estado', ''))
            ws.cell(row=row_num, column=18, value='Sí' if r.get('usa_ocs') else 'NO')
            ws.cell(row=row_num, column=19, value=r.get('observaciones', ''))
        
        # Ajustar ancho de columnas automáticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres de ancho
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta
        response = make_response(output.getvalue())
        filename = f'inventario_completo_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
        
    except Exception as e:
        flash(f'Error exportando datos: {e}', 'error')
        return redirect(url_for('index'))

# ==================== RUTAS DE IMPRESORAS ====================

@app.route('/impresoras')
def impresoras_dashboard():
    """Dashboard principal de impresoras"""
    try:
        impresoras = unified_web_db.get_all_printers()
        stats = unified_web_db.get_printers_statistics()
        return render_template('impresoras_dashboard.html', impresoras=impresoras, stats=stats)
    except Exception as e:
        flash(f'Error cargando impresoras: {e}', 'error')
        return render_template('impresoras_dashboard.html', impresoras=[], stats={})
    
@app.route('/impresoras/nueva', methods=['GET', 'POST'])
def impresoras_nueva():
    if request.method == 'GET':
        oficinas = unified_web_db.get_oficinas_list()
        marcas = unified_web_db.get_marcas_impresoras()
        return render_template('impresora_form.html', oficinas=oficinas, marcas=marcas, impresora=None, modo='nueva')
    else:
        try:
            # Obtener datos del formulario
            oficina_nombre = request.form.get('oficina', '')
            oficina_info = unified_web_db.get_oficina_with_piso(oficina_nombre)
            
            # ✅ IMPORTANTE: Obtener nombres de marca y modelo, no IDs
            marca_id = request.form.get('marca')
            modelo_id = request.form.get('modelo')
            
            # Buscar los nombres en la BD
            marca_nombre = unified_web_db.get_marca_nombre(marca_id)
            modelo_nombre = unified_web_db.get_modelo_nombre(modelo_id)
            
            data = {
                'oficina': oficina_nombre,
                'oficina_id': oficina_info['id'] if oficina_info else None,
                'piso': request.form.get('piso', ''),
                'marca': marca_nombre,  # ✅ Guardar nombre, no ID
                'modelo': modelo_nombre,  # ✅ Guardar nombre, no ID
                'numero_serie': request.form.get('numero_serie', ''),
                'tipo_conexion': request.form.get('tipo_conexion', 'red'),
                'ip': request.form.get('ip', ''),
                'toner': request.form.get('toner', ''),
                'observaciones': request.form.get('observaciones', ''),
                'estado': request.form.get('estado', 'Activa'),
                'usuario_registro': 'Web'
            }
            
            printer_id = unified_web_db.add_printer(data)
            
            if printer_id:
                flash('Impresora agregada exitosamente', 'success')
                return redirect(url_for('impresoras_dashboard'))
            else:
                flash('Error agregando impresora', 'error')
                oficinas = unified_web_db.get_oficinas_list()
                marcas = unified_web_db.get_marcas_impresoras()
                return render_template('impresora_form.html', oficinas=oficinas, marcas=marcas, impresora=None, modo='nueva')
                
        except Exception as e:
            flash(f'Error: {e}', 'error')
            oficinas = unified_web_db.get_oficinas_list()
            marcas = unified_web_db.get_marcas_impresoras()
            return render_template('impresora_form.html', oficinas=oficinas, marcas=marcas, impresora=None, modo='nueva')

@app.route('/impresoras/<int:printer_id>')
def impresoras_detalle(printer_id):
    """Ver detalle de una impresora"""
    try:
        impresora = unified_web_db.get_printer_by_id(printer_id)
        if not impresora:
            flash('Impresora no encontrada', 'error')
            return redirect(url_for('impresoras_dashboard'))
        
        return render_template('impresora_detalle.html', impresora=impresora)
    except Exception as e:
        flash(f'Error cargando detalle: {e}', 'error')
        return redirect(url_for('impresoras_dashboard'))
    

@app.route('/impresoras/<int:printer_id>/editar', methods=['GET', 'POST'])
def impresoras_editar(printer_id):
    if request.method == 'GET':
        try:
            impresora = unified_web_db.get_printer_by_id(printer_id)
            if not impresora:
                flash('Impresora no encontrada', 'error')
                return redirect(url_for('impresoras_dashboard'))
            
            oficinas = unified_web_db.get_oficinas_list()
            marcas = unified_web_db.get_marcas_impresoras()  # ✅ Agregar esto
            return render_template('impresora_editar.html', 
                                 impresora=impresora, 
                                 oficinas=oficinas,
                                 marcas=marcas)  # ✅ Pasar marcas      
        except Exception as e:
            flash(f'Error cargando impresora: {e}', 'error')
            return redirect(url_for('impresoras_dashboard'))
    else:
        try:
            # Obtener datos del formulario
            oficina_nombre = request.form.get('oficina', '')
            oficina_info = unified_web_db.get_oficina_with_piso(oficina_nombre)
            
            data = {
                'oficina': oficina_nombre,
                'oficina_id': oficina_info['id'] if oficina_info else None,
                'piso': request.form.get('piso', ''),
                'marca': request.form.get('marca', ''),
                'modelo': request.form.get('modelo', ''),
                'numero_serie': request.form.get('numero_serie', ''),
                'tipo_conexion': request.form.get('tipo_conexion', 'red'),
                'ip': request.form.get('ip', ''),
                'toner': request.form.get('toner', ''),
                'observaciones': request.form.get('observaciones', ''),
                'estado': request.form.get('estado', 'Activa')
            }
            
            success = unified_web_db.update_printer(printer_id, data)
            
            if success:
                flash('Impresora actualizada exitosamente', 'success')
            else:
                flash('Error actualizando impresora', 'error')
                
        except Exception as e:
            flash(f'Error: {e}', 'error')
        
        return redirect(url_for('impresoras_detalle', printer_id=printer_id))

@app.route('/impresoras/<int:printer_id>/eliminar', methods=['POST'])
def impresoras_eliminar(printer_id):
    """Eliminar una impresora"""
    try:
        success = unified_web_db.delete_printer(printer_id)
        if success:
            flash('Impresora eliminada exitosamente', 'success')
        else:
            flash('Error eliminando impresora', 'error')
    except Exception as e:
        flash(f'Error: {e}', 'error')
    
    return redirect(url_for('impresoras_dashboard'))

@app.route('/impresoras/estadisticas')
def impresoras_estadisticas():
    """Página de estadísticas de impresoras"""
    try:
        stats = unified_web_db.get_printers_statistics()
        impresoras = unified_web_db.get_all_printers()
        return render_template('impresoras_estadisticas.html', stats=stats, impresoras=impresoras)
    except Exception as e:
        flash(f'Error cargando estadísticas: {e}', 'error')
        return render_template('impresoras_estadisticas.html', stats={}, impresoras=[])

@app.route('/impresoras/exportar')
def impresoras_exportar():
    """Exportar impresoras a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        impresoras = unified_web_db.get_all_printers()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Impresoras"
        
        headers = ['ID', 'Oficina', 'Piso', 'Marca', 'Modelo', 'Número Serie', 
                  'Tipo Conexión', 'IP', 'Tóner', 'Estado', 'Observaciones', 'Fecha Registro']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for row_num, imp in enumerate(impresoras, 2):
            ws.cell(row=row_num, column=1, value=imp['id'])
            ws.cell(row=row_num, column=2, value=imp['oficina'])
            ws.cell(row=row_num, column=3, value=imp['piso'])
            ws.cell(row=row_num, column=4, value=imp['marca'])
            ws.cell(row=row_num, column=5, value=imp['modelo'])
            ws.cell(row=row_num, column=6, value=imp['numero_serie'])
            ws.cell(row=row_num, column=7, value=imp['tipo_conexion'])
            ws.cell(row=row_num, column=8, value=imp['ip'])
            ws.cell(row=row_num, column=9, value=imp['toner'])
            ws.cell(row=row_num, column=10, value=imp['estado'])
            ws.cell(row=row_num, column=11, value=imp['observaciones'])
            ws.cell(row=row_num, column=12, value=imp['fecha_registro'][:19] if imp['fecha_registro'] else '')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        filename = f'impresoras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
        
    except Exception as e:
        flash(f'Error exportando: {e}', 'error')
        return redirect(url_for('impresoras_dashboard'))
    
@app.route('/api/impresoras/modelos/<int:marca_id>')
def api_modelos_por_marca(marca_id):
    """API para obtener modelos según marca"""
    try:
        modelos = unified_web_db.get_modelos_por_marca(marca_id)
        return jsonify({'modelos': modelos})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/impresoras/toners/<int:modelo_id>')
def api_toners_por_modelo(modelo_id):
    """API para obtener toners según modelo"""
    try:
        toners = unified_web_db.get_toners_por_modelo(modelo_id)
        return jsonify({'toners': toners})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/impresoras/toner/<int:marca_id>/<int:modelo_id>')
def api_toner_especifico(marca_id, modelo_id):
    """API para obtener toner específico según marca y modelo"""
    try:
        toner = unified_web_db.get_toner_por_marca_modelo(marca_id, modelo_id)
        return jsonify({'toner': toner})
    except Exception as e:
        return jsonify({'error': str(e)}), 500    

# ==================== RUTAS DE NOTEBOOKS ====================

@app.route('/notebooks')
def notebooks_dashboard():
    """Dashboard principal de notebooks"""
    try:
        notebooks = unified_web_db.get_all_notebooks()
        
        # Estadísticas
        total = len(notebooks)
        prestadas = len([nb for nb in notebooks if nb['estado'] == 'Prestada'])
        guardadas = len([nb for nb in notebooks if nb['estado'] == 'Guardada'])
        
        stats = {
            'total': total,
            'prestadas': prestadas,
            'guardadas': guardadas
        }
        
        return render_template('notebooks_dashboard.html', notebooks=notebooks, stats=stats)
    except Exception as e:
        flash(f'Error cargando notebooks: {e}', 'error')
        return render_template('notebooks_dashboard.html', notebooks=[], stats={})

@app.route('/notebooks/agregar', methods=['GET', 'POST'])
def notebooks_agregar():
    """Formulario para agregar notebook desde inventario"""
    if request.method == 'GET':
        return render_template('notebooks_agregar.html')
    else:
        try:
            numero_serie = request.form.get('numero_serie', '').strip()
            
            if not numero_serie:
                flash('Debe ingresar un número de serie', 'error')
                return render_template('notebooks_agregar.html')
            
            exito, resultado = unified_web_db.trasladar_a_notebooks(numero_serie)
            
            if exito:
                flash(f'Notebook agregada exitosamente (ID: {resultado})', 'success')
                return redirect(url_for('notebooks_dashboard'))
            else:
                flash(f'Error: {resultado}', 'error')
                return render_template('notebooks_agregar.html', numero_serie=numero_serie)
                
        except Exception as e:
            flash(f'Error procesando solicitud: {e}', 'error')
            return render_template('notebooks_agregar.html')

@app.route('/notebooks/<int:notebook_id>')
def notebooks_detalle(notebook_id):
    """Ver detalle de una notebook específica"""
    try:
        notebook = unified_web_db.get_notebook_by_id(notebook_id)
        if not notebook:
            flash('Notebook no encontrada', 'error')
            return redirect(url_for('notebooks_dashboard'))
        
        historial = unified_web_db.get_notebook_historial(notebook_id)
        
        return render_template('notebooks_detalle.html', notebook=notebook, historial=historial)
    except Exception as e:
        flash(f'Error cargando detalle: {e}', 'error')
        return redirect(url_for('notebooks_dashboard'))

@app.route('/notebooks/<int:notebook_id>/prestar', methods=['POST'])
def notebooks_prestar(notebook_id):
    """Prestar una notebook"""
    try:
        persona = request.form.get('persona', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        
        if not persona:
            flash('Debe ingresar el nombre de la persona', 'error')
            return redirect(url_for('notebooks_detalle', notebook_id=notebook_id))
        
        if unified_web_db.prestar_notebook(notebook_id, persona, observaciones):
            flash(f'Notebook prestada a {persona} exitosamente', 'success')
        else:
            flash('Error al prestar notebook', 'error')
            
    except Exception as e:
        flash(f'Error: {e}', 'error')
    
    return redirect(url_for('notebooks_detalle', notebook_id=notebook_id))

@app.route('/notebooks/<int:notebook_id>/devolver', methods=['POST'])
def notebooks_devolver(notebook_id):
    """Devolver una notebook"""
    try:
        observaciones = request.form.get('observaciones', '').strip()
        
        if unified_web_db.devolver_notebook(notebook_id, observaciones):
            flash('Notebook devuelta exitosamente', 'success')
        else:
            flash('Error al devolver notebook', 'error')
            
    except Exception as e:
        flash(f'Error: {e}', 'error')
    
    return redirect(url_for('notebooks_detalle', notebook_id=notebook_id))

@app.route('/notebooks/<int:notebook_id>/cambiar_estado', methods=['POST'])
def notebooks_cambiar_estado(notebook_id):
    """Cambiar el estado de una notebook"""
    try:
        nuevo_estado = request.form.get('nuevo_estado', '').strip()
        observaciones = request.form.get('observaciones', '').strip()
        persona = request.form.get('persona', '').strip()
        
        if not nuevo_estado:
            flash('Estado no especificado', 'error')
            return redirect(url_for('notebooks_dashboard'))
        
        # Validar que si el estado es "Prestada", debe tener persona
        if nuevo_estado == 'Prestada' and not persona:
            flash('Debe asignar una persona cuando el estado es "Prestada"', 'error')
            return redirect(url_for('notebooks_dashboard'))
        
        if unified_web_db.cambiar_estado_notebook(notebook_id, nuevo_estado, observaciones, persona):
            flash(f'Notebook marcada como {nuevo_estado}', 'success')
        else:
            flash('Error al cambiar el estado', 'error')
            
    except Exception as e:
        flash(f'Error: {e}', 'error')
    
    return redirect(url_for('notebooks_dashboard'))

@app.route('/notebooks/exportar')
def notebooks_exportar():
    """Exportar notebooks a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        notebooks = unified_web_db.get_all_notebooks()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Notebooks"
        
        # Headers actualizados con "Nombre PC"
        headers = ['ID', 'Número de Serie', 'Nombre PC', 'Marca', 'Modelo', 'IP', 'MAC', 
                  'Estado', 'Persona Actual', 'Fecha Entrega', 'Fecha Devolución', 'Observaciones']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos de notebooks con pc_nombre incluido
        for row_num, nb in enumerate(notebooks, 2):
            ws.cell(row=row_num, column=1, value=nb['id'])
            ws.cell(row=row_num, column=2, value=nb['numero_serie'])
            ws.cell(row=row_num, column=3, value=nb.get('pc_nombre', ''))  # Nueva columna
            ws.cell(row=row_num, column=4, value=nb['marca'])
            ws.cell(row=row_num, column=5, value=nb['modelo'])
            ws.cell(row=row_num, column=6, value=nb['ip'])
            ws.cell(row=row_num, column=7, value=nb['mac_address'])
            ws.cell(row=row_num, column=8, value=nb['estado'])
            ws.cell(row=row_num, column=9, value=nb['persona_actual'])
            ws.cell(row=row_num, column=10, value=nb['fecha_entrega'][:19] if nb['fecha_entrega'] else '')
            ws.cell(row=row_num, column=11, value=nb['fecha_devolucion'][:19] if nb['fecha_devolucion'] else '')
            ws.cell(row=row_num, column=12, value=nb['observaciones'])
        
        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        filename = f'notebooks_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response
        
    except Exception as e:
        flash(f'Error exportando: {e}', 'error')
        return redirect(url_for('notebooks_dashboard'))
    
@app.route('/notebooks/<int:notebook_id>/editar', methods=['GET', 'POST'])
def notebooks_editar(notebook_id):
    """Editar información de una notebook"""
    if request.method == 'GET':
        try:
            notebook = unified_web_db.get_notebook_by_id(notebook_id)
            if not notebook:
                flash('Notebook no encontrada', 'error')
                return redirect(url_for('notebooks_dashboard'))
            
            return render_template('notebooks_editar.html', notebook=notebook)
        except Exception as e:
            flash(f'Error cargando notebook: {e}', 'error')
            return redirect(url_for('notebooks_dashboard'))
    
    else:  # POST
        try:
            # Obtener datos del formulario
            numero_serie = request.form.get('numero_serie', '').strip()
            pc_nombre = request.form.get('pc_nombre', '').strip()
            marca = request.form.get('marca', '').strip()
            modelo = request.form.get('modelo', '').strip()
            ip = request.form.get('ip', '').strip()
            mac_address = request.form.get('mac_address', '').strip()
            observaciones = request.form.get('observaciones', '').strip()
            estado = request.form.get('estado', '').strip()
            
            # Validar número de serie
            if not numero_serie:
                flash('El número de serie es obligatorio', 'error')
                return redirect(url_for('notebooks_editar', notebook_id=notebook_id))
            
            # Actualizar en la base de datos
            if unified_web_db.actualizar_notebook(
                notebook_id, 
                numero_serie, 
                pc_nombre, 
                marca, 
                modelo, 
                ip, 
                mac_address, 
                observaciones,
                estado
            ):
                flash('Notebook actualizada exitosamente', 'success')
                return redirect(url_for('notebooks_detalle', notebook_id=notebook_id))
            else:
                flash('Error al actualizar la notebook', 'error')
                return redirect(url_for('notebooks_editar', notebook_id=notebook_id))
                
        except Exception as e:
            flash(f'Error procesando solicitud: {e}', 'error')
            return redirect(url_for('notebooks_editar', notebook_id=notebook_id))
        
@app.route('/notebooks/<int:notebook_id>/borrar', methods=['POST'])
def notebooks_borrar(notebook_id):
    """Eliminar una notebook"""
    try:
        if unified_web_db.eliminar_notebook(notebook_id):
            flash('Notebook eliminada exitosamente', 'success')
        else:
            flash('Error al eliminar la notebook', 'error')
            
    except Exception as e:
        flash(f'Error: {e}', 'error')
    
    return redirect(url_for('notebooks_dashboard'))


# ==================== API ENDPOINTS ====================

@app.route('/api/oficina/<oficina_nombre>/piso')
def get_oficina_piso(oficina_nombre):
    """API endpoint para obtener el piso de una oficina"""
    try:
        oficina_info = unified_web_db.get_oficina_with_piso(oficina_nombre)
        if oficina_info:
            return jsonify({'piso': oficina_info['piso']})
        else:
            return jsonify({'piso': ''}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/buscar_mac/<mac_address>')
def buscar_mac_api(mac_address):
    """API endpoint para buscar registros por MAC address"""
    try:
        registros = unified_web_db.buscar_por_mac(mac_address.upper())
        return jsonify({
            'encontrados': len(registros),
            'registros': registros
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/macs_registradas')
def macs_registradas_api():
    """API endpoint para obtener todas las MACs registradas"""
    try:
        macs = unified_web_db.get_macs_registradas()
        return jsonify({
            'total': len(macs),
            'macs': macs
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/debug_bd')
def debug_bd():
    try:
        with unified_web_db.get_inventario_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(inventario)")
            columnas = cursor.fetchall()
            return f"Columnas en tabla inventario: {columnas}"
    except Exception as e:
        return f"Error: {e}"
@app.route('/debug_oficinas')
def debug_oficinas():
    try:
        with unified_web_db.get_oficinas_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, nombre_oficina, piso FROM oficinas ORDER BY id")
            oficinas = cursor.fetchall()
            return f"Oficinas en BD: {oficinas}"
    except Exception as e:
        return f"Error: {e}"

if __name__ == '__main__':
    app.run(
        host=app.config['HOST'],
        port=app.config['PORT'],
        debug=app.config['DEBUG']
    )

